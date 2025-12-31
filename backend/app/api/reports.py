"""
Reports API - Excel export
"""
import io
from datetime import datetime, date
from decimal import Decimal
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import joinedload
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from typing import Optional, List

from app.database import get_db
from app.utils.deps import get_admin_user
from app.models.user import User
from app.models.transaction import Transaction, TransactionType

router = APIRouter()


@router.get("/export")
async def export_transactions(
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    agent_ids: Optional[List[int]] = Query(None),
    types: Optional[List[str]] = Query(None),
    admin: User = Depends(get_admin_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Export transactions to Excel file.

    Filters:
    - date_from, date_to: Date range
    - agent_ids: Filter by specific agent user IDs
    - types: Transaction types (CHARGE_PENDING, CHARGE_APPROVED, etc.)
    """
    query = (
        select(Transaction)
        .options(joinedload(Transaction.user))
        .order_by(Transaction.created_at.desc())
    )

    # Apply filters
    if date_from:
        query = query.where(Transaction.created_at >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        query = query.where(Transaction.created_at <= datetime.combine(date_to, datetime.max.time()))
    if agent_ids:
        query = query.where(Transaction.user_id.in_(agent_ids))
    if types:
        type_enums = [TransactionType(t) for t in types if t in [e.value for e in TransactionType]]
        if type_enums:
            query = query.where(Transaction.type.in_(type_enums))

    result = await db.execute(query)
    transactions = result.scalars().all()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        "ID",
        "تاریخ",
        "ساعت",
        "کاربر",
        "نوع تراکنش",
        "مبلغ",
        "موجودی قبل",
        "موجودی بعد",
        "توضیحات"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Data rows
    for row_num, tx in enumerate(transactions, 2):
        ws.cell(row=row_num, column=1, value=tx.id)
        ws.cell(row=row_num, column=2, value=tx.created_at.strftime("%Y-%m-%d"))
        ws.cell(row=row_num, column=3, value=tx.created_at.strftime("%H:%M:%S"))
        ws.cell(row=row_num, column=4, value=tx.user.username if tx.user else "")
        ws.cell(row=row_num, column=5, value=tx.type.value)
        ws.cell(row=row_num, column=6, value=float(tx.amount))
        ws.cell(row=row_num, column=7, value=float(tx.balance_before))
        ws.cell(row=row_num, column=8, value=float(tx.balance_after))
        ws.cell(row=row_num, column=9, value=tx.notes or "")

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Generate filename
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"RAD_Report_{now}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/stats")
async def get_stats(
    admin: User = Depends(get_admin_user),
    db: AsyncSession = Depends(get_db)
):
    """Get dashboard statistics (Admin only)"""
    from sqlalchemy import func
    from app.models.agent import Agent
    from app.models.order import Order, OrderStatus
    from app.models.payment import Payment, PaymentStatus

    # Count agents
    result = await db.execute(select(func.count(Agent.id)))
    total_agents = result.scalar() or 0

    # Count active orders
    result = await db.execute(
        select(func.count(Order.id)).where(Order.status == OrderStatus.ACTIVE)
    )
    active_orders = result.scalar() or 0

    # Count pending payments
    result = await db.execute(
        select(func.count(Payment.id)).where(Payment.status == PaymentStatus.PENDING)
    )
    pending_payments = result.scalar() or 0

    # Total revenue (approved payments)
    result = await db.execute(
        select(func.sum(Payment.amount)).where(Payment.status == PaymentStatus.APPROVED)
    )
    total_revenue = result.scalar() or Decimal(0)

    return {
        "total_agents": total_agents,
        "active_orders": active_orders,
        "pending_payments": pending_payments,
        "total_revenue": float(total_revenue)
    }
